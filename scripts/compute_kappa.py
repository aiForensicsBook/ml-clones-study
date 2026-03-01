#!/usr/bin/env python3

import argparse
import sys
from collections import Counter

try:
    from openpyxl import load_workbook
except ImportError:
    print("pip install openpyxl --break-system-packages")
    sys.exit(1)

def read_classifications(filepath, sheet_name):
    wb = load_workbook(filepath, data_only=True)
    ws = wb[sheet_name]

    classifications = []
    for row in ws.iter_rows(min_row=5, values_only=False):
        num_cell = row[0]
        cls_cell = row[7]
        name_cell = row[3]
        proj_cell = row[1]

        if num_cell.value is None:
            break

        cls = str(cls_cell.value).strip().upper() if cls_cell.value else None
        if cls not in ("A", "B", "C"):
            print(f"  Warning: Row {num_cell.value} has invalid classification '{cls_cell.value}', skipping")
            continue

        classifications.append({
            "num": num_cell.value,
            "project": proj_cell.value,
            "name": name_cell.value,
            "classification": cls,
        })

    return classifications

def cohens_kappa(r1_labels, r2_labels, categories=("A", "B", "C")):
    n = len(r1_labels)
    assert n == len(r2_labels), "Label lists must be same length"

    agree = sum(1 for a, b in zip(r1_labels, r2_labels) if a == b)
    po = agree / n

    pe = 0
    for cat in categories:
        p1 = sum(1 for x in r1_labels if x == cat) / n
        p2 = sum(1 for x in r2_labels if x == cat) / n
        pe += p1 * p2

    if pe == 1.0:
        return 1.0

    kappa = (po - pe) / (1 - pe)
    return kappa

def weighted_kappa(r1_labels, r2_labels, categories=("A", "B", "C")):
    n = len(r1_labels)
    k = len(categories)
    cat_idx = {c: i for i, c in enumerate(categories)}

    observed = [[0] * k for _ in range(k)]
    for a, b in zip(r1_labels, r2_labels):
        observed[cat_idx[a]][cat_idx[b]] += 1

    row_totals = [sum(observed[i]) for i in range(k)]
    col_totals = [sum(observed[i][j] for i in range(k)) for j in range(k)]
    expected = [[row_totals[i] * col_totals[j] / n for j in range(k)] for i in range(k)]

    weights = [[1 - abs(i - j) / (k - 1) for j in range(k)] for i in range(k)]

    po_w = sum(weights[i][j] * observed[i][j] / n for i in range(k) for j in range(k))
    pe_w = sum(weights[i][j] * expected[i][j] / n for i in range(k) for j in range(k))

    if pe_w == 1.0:
        return 1.0

    return (po_w - pe_w) / (1 - pe_w)

def confusion_matrix(r1_labels, r2_labels, categories=("A", "B", "C")):
    k = len(categories)
    cat_idx = {c: i for i, c in enumerate(categories)}
    matrix = [[0] * k for _ in range(k)]
    for a, b in zip(r1_labels, r2_labels):
        matrix[cat_idx[a]][cat_idx[b]] += 1
    return matrix

def main():
    parser = argparse.ArgumentParser(description="Compute Cohen's kappa for intra-rater validation")
    parser.add_argument("--round1", required=True, help="Path to Round 1 Excel workbook")
    parser.add_argument("--round2", required=True, help="Path to Round 2 Excel workbook")
    parser.add_argument("--output", default="results/kappa_results.csv", help="Output CSV path")
    args = parser.parse_args()

    print("=" * 60)
    print("INTRA-RATER VALIDATION: COHEN'S KAPPA")
    print("=" * 60)

    print(f"\nReading Round 1: {args.round1}")
    r1 = read_classifications(args.round1, "Round 1")
    print(f"  Found {len(r1)} classifications")

    print(f"Reading Round 2: {args.round2}")
    r2 = read_classifications(args.round2, "Round 2")
    print(f"  Found {len(r2)} classifications")

    r2_lookup = {}
    for item in r2:
        key = (item["project"], item["name"])
        r2_lookup[key] = item["classification"]

    matched_r1 = []
    matched_r2 = []
    unmatched = 0
    for item in r1:
        key = (item["project"], item["name"])
        if key in r2_lookup:
            matched_r1.append(item["classification"])
            matched_r2.append(r2_lookup[key])
        else:
            unmatched += 1
            print(f"  Warning: No Round 2 match for {key}")

    if unmatched > 0:
        print(f"\n  {unmatched} units could not be matched between rounds")

    n = len(matched_r1)
    print(f"\n  Matched pairs: {n}")

    if n == 0:
        print("ERROR: No matched pairs found. Check workbook format.")
        sys.exit(1)

    categories = ("A", "B", "C")
    k = cohens_kappa(matched_r1, matched_r2, categories)
    k_w = weighted_kappa(matched_r1, matched_r2, categories)
    cm = confusion_matrix(matched_r1, matched_r2, categories)

    agree = sum(1 for a, b in zip(matched_r1, matched_r2) if a == b)
    pct_agree = 100 * agree / n

    print(f"\n{'='*60}")
    print(f"RESULTS")
    print(f"{'='*60}")
    print(f"  Sample size:              {n}")
    print(f"  Raw agreement:            {agree}/{n} ({pct_agree:.1f}%)")
    print(f"  Cohen's kappa:            {k:.4f}")
    print(f"  Weighted kappa (linear):  {k_w:.4f}")

    if k >= 0.81:
        interp = "Almost perfect agreement"
    elif k >= 0.61:
        interp = "Substantial agreement"
    elif k >= 0.41:
        interp = "Moderate agreement"
    elif k >= 0.21:
        interp = "Fair agreement"
    else:
        interp = "Slight or poor agreement"
    print(f"  Interpretation:           {interp} (Landis & Koch, 1977)")

    print(f"\n  Confusion Matrix (rows=Round 1, cols=Round 2):")
    print(f"       {'  '.join(categories)}")
    for i, cat in enumerate(categories):
        row_str = "  ".join(f"{cm[i][j]:3d}" for j in range(len(categories)))
        print(f"  {cat}:  {row_str}")

    print(f"\n  Per-category agreement:")
    for i, cat in enumerate(categories):
        total_cat = sum(cm[i])
        if total_cat > 0:
            cat_agree = cm[i][i]
            print(f"    {cat}: {cat_agree}/{total_cat} ({100*cat_agree/total_cat:.1f}%)")

    print(f"\n  Disagreement pairs:")
    disagree_counter = Counter()
    for a, b in zip(matched_r1, matched_r2):
        if a != b:
            disagree_counter[f"{a}->{b}"] += 1
    for pair, count in disagree_counter.most_common():
        print(f"    {pair}: {count}")

    import csv
    with open(args.output, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["Metric", "Value"])
        writer.writerow(["Sample size", n])
        writer.writerow(["Raw agreement", f"{agree}/{n} ({pct_agree:.1f}%)"])
        writer.writerow(["Cohen's kappa", f"{k:.4f}"])
        writer.writerow(["Weighted kappa", f"{k_w:.4f}"])
        writer.writerow(["Interpretation", interp])
        writer.writerow([])
        writer.writerow(["Confusion Matrix", ""] + list(categories))
        for i, cat in enumerate(categories):
            writer.writerow([cat] + [""] + cm[i])

    print(f"\n  Results saved to: {args.output}")
    print(f"\n{'='*60}")
    print(f"VALIDATION COMPLETE")
    print(f"{'='*60}")

if __name__ == "__main__":
    main()
