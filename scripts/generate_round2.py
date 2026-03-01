#!/usr/bin/env python3

import json
import random
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

random.seed(99)

HEADER_FONT = Font(name='Arial', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill('solid', fgColor='7030A0')
DATA_FONT = Font(name='Arial', size=10)
TITLE_FONT = Font(name='Arial', bold=True, size=14, color='7030A0')
YELLOW_FILL = PatternFill('solid', fgColor='FFF9C4')
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

meta_path = Path("results/validation_sample.json")
if not meta_path.exists():
    print(f"ERROR: {meta_path} not found. Run build_validation.py first.")
    exit(1)

with open(meta_path) as f:
    meta = json.load(f)

units = meta["units"]
random.shuffle(units)

wb = Workbook()
ws = wb.active
ws.title = "Round 2"
ws.sheet_properties.tabColor = "7030A0"

ws["A1"] = "Round 2 — Manual Classification (14-day gap)"
ws["A1"].font = TITLE_FONT
ws.merge_cells("A1:H1")
ws["A2"] = "Date completed: _______________    Rater: Joseph C. Sremack"
ws["A2"].font = Font(name="Arial", italic=True, size=10)
ws.merge_cells("A2:H2")
ws["A3"] = "Classify each unit as A (Boilerplate), B (API Protocol), or C (Custom). See Round 1 Instructions sheet for definitions."
ws["A3"].font = DATA_FONT
ws.merge_cells("A3:H3")

headers = ["#", "Project", "File", "Unit Name", "Type", "Lines",
           "Code Snippet", "Your Classification", "Notes"]
for col_idx, h in enumerate(headers, 1):
    c = ws.cell(row=5, column=col_idx, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal="center", wrap_text=True)
    c.border = THIN_BORDER

dv = DataValidation(type="list", formula1='"A,B,C"', allow_blank=True)
dv.error = "Please select A, B, or C"
ws.add_data_validation(dv)

for i, unit in enumerate(units):
    r = 6 + i
    loc = unit["line_end"] - unit["line_start"] + 1

    ws.cell(row=r, column=1, value=i + 1).font = DATA_FONT
    ws.cell(row=r, column=1).alignment = Alignment(horizontal="center")
    ws.cell(row=r, column=1).border = THIN_BORDER

    ws.cell(row=r, column=2, value=unit["project"]).font = DATA_FONT
    ws.cell(row=r, column=2).border = THIN_BORDER

    ws.cell(row=r, column=3, value=unit["file"]).font = DATA_FONT
    ws.cell(row=r, column=3).border = THIN_BORDER

    ws.cell(row=r, column=4, value=unit["name"]).font = DATA_FONT
    ws.cell(row=r, column=4).border = THIN_BORDER

    ws.cell(row=r, column=5, value=unit["type"]).font = DATA_FONT
    ws.cell(row=r, column=5).alignment = Alignment(horizontal="center")
    ws.cell(row=r, column=5).border = THIN_BORDER

    ws.cell(row=r, column=6, value=f"L{unit['line_start']}-{unit['line_end']} ({loc})").font = DATA_FONT
    ws.cell(row=r, column=6).alignment = Alignment(horizontal="center")
    ws.cell(row=r, column=6).border = THIN_BORDER

    ws.cell(row=r, column=7, value=unit["snippet"]).font = Font(name="Consolas", size=9)
    ws.cell(row=r, column=7).alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(row=r, column=7).border = THIN_BORDER

    cls_cell = ws.cell(row=r, column=8)
    cls_cell.font = Font(name="Arial", bold=True, size=12)
    cls_cell.alignment = Alignment(horizontal="center", vertical="center")
    cls_cell.border = THIN_BORDER
    cls_cell.fill = YELLOW_FILL
    dv.add(cls_cell)

    ws.cell(row=r, column=9).font = DATA_FONT
    ws.cell(row=r, column=9).border = THIN_BORDER

    ws.row_dimensions[r].height = min(max(loc * 13, 40), 300)

ws.freeze_panes = "A6"
ws.auto_filter.ref = f"A5:I{5 + len(units)}"

for i, w in enumerate([4, 22, 25, 28, 8, 14, 70, 18, 25], 1):
    ws.column_dimensions[get_column_letter(i)].width = w

OUT = "results/intra_rater_round2.xlsx"
wb.save(OUT)
print(f"Round 2 workbook saved to: {OUT}")
print(f"Units: {len(units)}, reshuffled with seed=99")
print(f"\nAfter completing both rounds, run:")
print(f"  python scripts/compute_kappa.py --round1 results/intra_rater_round1.xlsx --round2 results/intra_rater_round2.xlsx")
